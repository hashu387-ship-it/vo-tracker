#!/usr/bin/env python3
"""
Extract the HW2C05 commercial register out of the source Excel workbook
("Pay Reg & VO LOG HW2 MEP") into the JSON seed files under data/seed/.

Usage:
    python3 scripts/extract-workbook.py <path-to-workbook.xlsx>

The parser mirrors lib/excel/import.ts so a re-import through the web UI
produces exactly the same records as the checked-in seed.

Source layout (as of the 12-May-2026 revision):
  sheet "VO LOG (2)"       header on row 10, records from row 11
  sheet "Payment Register" header on row 15, records from row 17
"""
from __future__ import annotations

import datetime as _dt
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "seed"

VO_SHEET = "VO LOG (2)"
VO_FIRST_ROW = 11
PAY_SHEET = "Payment Register"
PAY_FIRST_ROW = 17
PAY_LAST_ROW = 49


def text(value):
    if value is None:
        return None
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    cleaned = re.sub(r"\s+", " ", str(value)).strip()
    return cleaned or None


def number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(",", "").replace("SAR", "").strip()), 2)
    except ValueError:
        return None


def date(value):
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", value.strip())
        if match:
            return f"{match.group(3)}-{int(match.group(2)):02d}-{int(match.group(1)):02d}"
    return None


def extract_variations(wb):
    ws = wb[VO_SHEET]

    def link(row: int, column: str):
        cell = ws[f"{column}{row}"]
        return cell.hyperlink.target if cell.hyperlink else None

    rows = []
    for r in range(VO_FIRST_ROW, ws.max_row + 1):
        serial = ws.cell(r, 2).value
        if not isinstance(serial, (int, float)):
            # Totals rows and free-text notes at the bottom of the sheet.
            continue
        rows.append(
            {
                "serial": int(serial),
                "voNumber": text(ws.cell(r, 3).value),
                "aconexDate": date(ws.cell(r, 4).value),
                "dvoReference": text(ws.cell(r, 5).value),
                "subject": text(ws.cell(r, 6).value),
                "submissionDate": date(ws.cell(r, 7).value),
                "submissionType": text(ws.cell(r, 8).value),
                "submissionRef": text(ws.cell(r, 9).value),
                "responseRef": text(ws.cell(r, 10).value),
                "proposalValue": number(ws.cell(r, 12).value),
                "clientAssessment": number(ws.cell(r, 13).value),
                "agreedValue": number(ws.cell(r, 14).value),
                "status": text(ws.cell(r, 15).value),
                "vorRef": text(ws.cell(r, 16).value),
                "dvoRef": text(ws.cell(r, 17).value),
                "contractorRemarks": text(ws.cell(r, 18).value),
                "clientRemarks": text(ws.cell(r, 19).value),
                "aconexLink": link(r, "U"),
                "submissionLink": link(r, "V"),
                "owner": text(ws.cell(r, 26).value),
            }
        )
    return rows


def extract_payments(wb):
    ws = wb[PAY_SHEET]
    rows = []
    for r in range(PAY_FIRST_ROW, PAY_LAST_ROW + 1):
        ref = text(ws.cell(r, 1).value)
        if not ref:
            continue
        rows.append(
            {
                "ref": ref,
                "period": text(ws.cell(r, 2).value),
                "grossCertified": number(ws.cell(r, 4).value),
                "advanceRecovery": number(ws.cell(r, 5).value),
                "backCharge": number(ws.cell(r, 6).value),
                "retention": number(ws.cell(r, 7).value),
                "vatOnAdvanceRecovery": number(ws.cell(r, 8).value),
                "vat": number(ws.cell(r, 9).value),
                "netCertified": number(ws.cell(r, 10).value),
                "received": number(ws.cell(r, 11).value),
                "submittedDate": date(ws.cell(r, 13).value),
                "taxInvoiceDate": date(ws.cell(r, 14).value),
                "dueDate": date(ws.cell(r, 15).value),
                "paymentNote": text(ws.cell(r, 16).value),
                "status": text(ws.cell(r, 17).value),
                "collectedDate": date(ws.cell(r, 18).value),
                "contractorAction": text(ws.cell(r, 19).value),
                "clientAction": text(ws.cell(r, 20).value),
                "cumulativeGross": number(ws.cell(r, 21).value),
            }
        )
    return rows


def extract_project(wb):
    ws = wb[PAY_SHEET]
    return {
        "code": text(ws["B4"].value),
        "name": "Shura West Hotel 02 — MEP Package",
        "contractor": text(ws["B5"].value),
        "client": "Red Sea Global",
        "contractDate": date(ws["B6"].value),
        "currency": "SAR",
        "originalContractValue": number(ws["D4"].value),
        "revisedContractValue": number(ws["D5"].value),
        "advancePaymentTotal": number(ws["H4"].value),
        "advancePaymentPercent": 0.30,
        "retentionCapPercent": 0.05,
        "vatRate": 0.15,
        "dataAsOf": date(wb[VO_SHEET]["B1"].value),
        "sourceWorkbook": "Pay Reg & VO LOG HW2 MEP 12May2026.xlsx",
    }


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    OUT.mkdir(parents=True, exist_ok=True)
    project = extract_project(wb)
    variations = extract_variations(wb)
    payments = extract_payments(wb)

    (OUT / "project.json").write_text(json.dumps(project, indent=2) + "\n")
    (OUT / "variations.json").write_text(json.dumps(variations, indent=2) + "\n")
    (OUT / "payments.json").write_text(json.dumps(payments, indent=2) + "\n")

    print(f"project        → {OUT/'project.json'}")
    print(f"{len(variations):>3} variations → {OUT/'variations.json'}")
    print(f"{len(payments):>3} payments   → {OUT/'payments.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
